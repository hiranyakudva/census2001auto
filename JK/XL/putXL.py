import openpyxl
max=14
jk = openpyxl.load_workbook('JK.xlsx')
jk_sheet = jk.get_sheet_by_name('Sheet1')
jk.get_active_sheet()

for i in range(1,max+1):

	if (i<10):
		string='0'+str(i)
	else:
		string=str(i)
	xl_file_name="output"+string+".xlsx"

	district=openpyxl.load_workbook(xl_file_name)
	district_sheet=district.get_sheet_by_name('Page 1')
	district.get_active_sheet()

	#district name
	jk_sheet.cell(row=i,column=0).value=district_sheet.cell(row=1,column=1).value

	#persons to SC % Vertically Down
	j=1
	a=4
	
	while (j<9):
		#jk_sheet.cell(row=i,column=j).value=1 ---WORKING FINE
		#test=district_sheet(row=a,column=1) --- TYPE ERROR
		#test=district_sheet['A2'] --- WORKING FINE
		#jk_sheet.cell(row=i,column=j).value=test --- WORKS IF TEST IS DEFINED CORRECTLY
		#jk_sheet.cell(row=i,column=j).value=district_sheet(row=a,column=1) --- TYPE ERROR 
		a=a+1
		j=j+1

	#nHouseholds
	jk_sheet.cell(row=i,column=9).value=district_sheet(row=4,column=3).value
	#houseHoldSize
	jk_sheet.cell(row=i,column=10).value=district_sheet(row=5,column=3).value
	#sexRation
	jk_sheet.cell(row=i,column=11).value=district_sheet(row=7,column=3).value
	#sexRatio(0-6)
	jk_sheet.cell(row=i,column=12).value=district_sheet(row=8,column=3).value
	#ST
	jk_sheet.cell(row=i,column=13).value=district_sheet(row=10,column=3).value
	#%ST
	jk_sheet.cell(row=i,column=14).value=district_sheet(row=11,column=3).value

	#literates
	jk_sheet.cell(row=i,column=15).value=district_sheet(row=14,column=1).value
	jk_sheet.cell(row=i,column=16).value=district_sheet(row=15,column=1).value
	jk_sheet.cell(row=i,column=17).value=district_sheet(row=16,column=1).value

	#literacy_rate
	jk_sheet.cell(row=i,column=18).value=district_sheet(row=18,column=1).value
	jk_sheet.cell(row=i,column=19).value=district_sheet(row=19,column=1).value
	jk_sheet.cell(row=i,column=20).value=district_sheet(row=20,column=1).value

	#education_whole
	j=21
	a=14
	while (j<28):
		jk_sheet.cell(row=i,column=j).value=district_sheet(row=a,column=3).value
		j=j+1
		a=a+1

	#age_groups
	j=28
	a=22
	while (j<32):
		jk_sheet.cell(row=i,column=j).value=district_sheet(row=a,column=3).value
		j=j+1
		a=a+1

	#workers
	j=32
	a=22
	while (j<36):
		jk_sheet.cell(row=i,column=j).value=district_sheet(row=a,column=1).value
		j=j+1
		a=a+1

	#sc_name
	jk_sheet.cell(row=i,column=36).value=district_sheet(row=27,column=0).value
	jk_sheet.cell(row=i,column=38).value=district_sheet(row=28,column=0).value
	jk_sheet.cell(row=i,column=40).value=district_sheet(row=29,column=0).value

	#sc_pop
	jk_sheet.cell(row=i,column=37).value=district_sheet(row=27,column=1).value
	jk_sheet.cell(row=i,column=39).value=district_sheet(row=28,column=1).value
	jk_sheet.cell(row=i,column=41).value=district_sheet(row=29,column=1).value

	#religeon_name
	jk_sheet.cell(row=i,column=42).value=district_sheet(row=31,column=0).value
	jk_sheet.cell(row=i,column=44).value=district_sheet(row=32,column=0).value
	jk_sheet.cell(row=i,column=46).value=district_sheet(row=33,column=0).value

	#religeon_pop
	jk_sheet.cell(row=i,column=43).value=district_sheet(row=31,column=1).value
	jk_sheet.cell(row=i,column=45).value=district_sheet(row=32,column=1).value
	jk_sheet.cell(row=i,column=47).value=district_sheet(row=33,column=1).value

	#st_name
	jk_sheet.cell(row=i,column=48).value=district_sheet(row=27,column=2).value
	jk_sheet.cell(row=i,column=50).value=district_sheet(row=28,column=2).value
	jk_sheet.cell(row=i,column=52).value=district_sheet(row=29,column=2).value	

	#st_pop
	jk_sheet.cell(row=i,column=49).value=district_sheet(row=27,column=3).value
	jk_sheet.cell(row=i,column=51).value=district_sheet(row=28,column=3).value
	jk_sheet.cell(row=i,column=53).value=district_sheet(row=29,column=3).value

	#imp_towns_names
	jk_sheet.cell(row=i,column=54).value=district_sheet(row=39,column=0).value
	jk_sheet.cell(row=i,column=56).value=district_sheet(row=40,column=0).value
	jk_sheet.cell(row=i,column=58).value=district_sheet(row=41,column=0).value

	#imp_towns_pop
	jk_sheet.cell(row=i,column=55).value=district_sheet(row=39,column=1).value
	jk_sheet.cell(row=i,column=57).value=district_sheet(row=40,column=1).value
	jk_sheet.cell(row=i,column=59).value=district_sheet(row=41,column=1).value

	#inhabited_villages
	jk_sheet.cell(row=i,column=60).value=district_sheet(row=31,column=3).value

	#Ameneties
	j=61
	a=35
	while (j<77):
		jk_sheet.cell(row=i,column=j).value=district_sheet(row=a,column=3).value
		j=j+1
		a=a+1

	#houses
	j=77
	a=48
	while (j<80):
		jk_sheet.cell(row=i,column=j).value=district_sheet(row=a,column=1).value
		j=j+1
		a=a+1

jk.save("final.xlsx")




